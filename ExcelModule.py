import re
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from CellData import Order
from CellData import Category


class Spreadsheet(object):
    #each line is a category to find
    #make different regex algorithms based on spreadsheet type
    regex_category = [] #stores the names of each category

    regex = [] #stores each regex category individually
    regex.append( re.compile(r'^(\s* first \s* name \s*)$', re.I | re.X) )
    regex_category.append('first name')

    regex.append( re.compile(r'^(\s* last \s* name \s*)$', re.I | re.X) )
    regex_category.append('last name')

    regex.append( re.compile(r'^(\s* company \s*)$', re.I | re.X) )
    regex_category.append('company')

    regex.append( re.compile(r'^(\s* address \s* line \s* \d*)$', re.I | re.X) )
    regex_category.append('address line')

    regex.append( re.compile(r'^(\s* city \s*)$', re.I | re.X) )
    regex_category.append('city')

    regex.append( re.compile(r'^(\s* state \s*)$', re.I | re.X) )
    regex_category.append('state')

    regex.append( re.compile(r'^(\s* zip \s*)$', re.I | re.X) )
    regex_category.append('zip')

    regex.append( re.compile(r'^(\s* email \s*)$', re.I | re.X) )
    regex_category.append('email')

    regex.append( re.compile(r'^(\s* (qty | quantity) \s* \d* \s*)$', re.I | re.X) )
    regex_category.append('quantity')

    regex.append( re.compile(r'^(\s* product \s* \d* \s*)$', re.I | re.X) )
    regex_category.append('product')

    regex.append( re.compile(r'''^(\s* product (\s*|\w*) \d* \s*
                              (qty | quantity) (\s* | \w*) \d* \s*)$''', re.I | re.X) )
    regex_category.append('product and quantity')

    def __init__(self, wb_string):
        try:
            self.wb = openpyxl.load_workbook(wb_string) #current workbook
            self.ws = self.wb.active #current sheet from workbook
            self.total_rows = len( list(self.ws.rows) ) #total rows for current spreadsheet
            self.total_cols = len( list(self.ws.columns) ) #total cols for current spreadsheet
            self.category_row = -1 #starting row for categories.
            self.num_of_categories = -1 #number of categories found
            self.order_categories = {} #holds all categories w/row&column position
            self.orders = [] #list of all orders for current spreadsheet

        except Exception as e:
            print('{0}'.format(e) )


    def is_cell_category(self, row, col): #may need to return index 'i' to know what category is associated with which column
        cell_str = self.ws.cell(row = row, column = col).value

        if cell_str is not None: #cell has a value
            for i in range( len(Spreadsheet.regex) ):
                if Spreadsheet.regex[i].fullmatch( str(cell_str) ) is not None:
                    return i #index of category found

            return None

    #finds all the categories. This goes to O(n) from finding all cells in sheet
    def get_all_categories(self):
        cell_str = '' #holds cell value
        category_index = '' #determines which category was read in

        for col in range(1, self.total_cols + 1): #goes over category columns
            cell_str = str(self.ws.cell(row = self.category_row,
                                        column = col).value)

            if cell_str is not None: #cell has a value
                category_index = self.is_cell_category(self.category_row, col)
                if category_index is not None: #cell is a category
                    category_obj = Category(Spreadsheet.regex_category[category_index], self.category_row, col) #stores current row & column
                    cell_pos = '{0}'.format(get_column_letter(col))
                    #format category column letter : category object
                    self.order_categories[cell_pos] = category_obj

        self.num_of_categories = len( self.order_categories )
        print('number of categories:' + str(self.num_of_categories) )
        return None


    #finds the categories on the spreadsheet using regex
    def find_category_row(self):
        search = '' #used to store the found category

        for row in range(1, self.total_rows + 1):
            for col in range(1, self.total_cols + 1):
                search = self.is_cell_category(row, col) #looks for a category in the current position

                if search is not None: #category found
                    if self.category_row == -1: #category row not assigned
                        self.category_row = row #first category row found
                        break

        return self.get_all_categories() #finds remaining categories w/category row


    def get_orders(self): #called after categories are found
        current_order = Order()
        cell_val = ''
        val_exists = ''
        current_product = None
        current_qty = None

        #goes through all the rows, saving the orders
        for row in range(self.category_row + 1, self.total_rows + 1):
            val_exists = False
            current_product = None
            current_qty = None
            current_order = Order() #reset

            #for key, val in self.order_categories.items(): #goes through each category for the order
            for col in range(1, self.num_of_categories + 2): #+2 accounts for cells starting at 1, and range() not going to the upper bound
                cell_val = self.ws.cell(row = row, column = col).value

                if cell_val is not None: #the cell has a value
                    if get_column_letter(col) in self.order_categories.keys(): #if the key exists
                        if self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[0]:
                            current_order.first_name = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[1]:
                            current_order.last_name = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[2]:
                            current_order.company = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[3]:
                            if current_order.address_1 is None:
                                current_order.address_1 = cell_val

                            else:
                                current_order.address_2 = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[4]:
                            current_order.city = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[5]:
                            current_order.state = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[6]:
                            current_order.zip = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[7]:
                            current_order.email = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[8]:
                            current_qty = cell_val

                        elif self.order_categories[ get_column_letter(col) ].name == Spreadsheet.regex_category[9]:
                            current_product = cell_val

                        if current_product is not None and current_qty is not None:
                            current_order.product_and_qty[current_product] = current_qty
                            print('Product: {0}  Qty: {1} (Added)'.format(current_product, current_qty) )
                            current_product = None #reset
                            current_qty = None #reset

                        val_exists = True

            if val_exists:
                self.orders.append(current_order)
                print()

        return None

    def print_orders(self):
        for i in range( len(self.orders) ):
            ending = '\n' + ('_' * 40) + '\n\n'
            print( self.orders[i], end= ending)
