import re
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter
from CellData import Order
from ExcelModule import Spreadsheet


def find_all(list_match):
    print('{} {}'.format('All matches:', list_match), end='\n\n' )

    for match in list_match:
        if type(match) == tuple:
            print('{}'.format(''.join(match) ), end='\n\n')

        else:
            print('{}'.format(match), end='\n\n')

    return None

def main():
    #to_search = 'asfdExample:37Da457sdfuck f7uagdc45368k asdffgExamplesohASD84t' #print Example/Examples/fuck
    #to_search = 'asldkjf4350934asldf(234) 523-5111, as(111)-222.3333lkdjf254-5555!:@555 . 123.1111$)($%)' #print the phone numbers
    #to_search = '12 drummers, 11 whatever, 10 whats, 9 off, 8 fucks, 7 done' #
    #to_search = 'agent smith gave AGent Cooper documents' #sensor the agent name

    '''
    for row in ws_data.iter_rows():
        blank_row = True
        for cell in row:
            if cell.value is not None:
                print( '{0}: {1} {2}'.format( type(cell.value), cell, cell.value ) )
                blank_row = False

        if not blank_row:
            print(end='{0} {1}'.format( ('_' * 60), '\n\n') )


    wb = openpyxl.load_workbook('spring_mobile.xlsx')
    ws = wb.active #active worksheet

    Order().find_category_row(ws)
    '''

    regex = re.compile(r'''
                      ^(\s* first \s* name \s*)$
                      | ^(\s* last \s* name \s*)$
                      | ^(\s* company \s*)$
                      | ^(\s* address \s* line \s* \d*)$ #gets both addresses
                      | ^(\s* city \s*)$
                      | ^(\s* state \s*)$
                      | ^(\s* zip \s*)$
                      | ^(\s* email \s*)$
                      | ^(\s* qty \s* \d* \s*)$
                      | ^(\s* product \s* \d* \s*)$

                         #matches product and quantity
                      | ^(\s* product (\s*|\w*) \d* \s*
                         (qty | quantity) (\s* | \w*) \d* \s*)$
                      ''', re.X | re.I | re.M)

    match = regex.fullmatch('email')
    print( match.group(8) ) #use the match to determine what was read in(for storage)

    return None

main()
